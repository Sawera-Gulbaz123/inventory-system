# backend/routes/backup.py

import os
import io
import subprocess
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List

import models
import schemas
from database import get_db, DATABASE_URL
from auth_utils import require_admin, get_current_user, log_activity

router = APIRouter(prefix="/backup", tags=["Backup"])


# ── HELPER: Parse DB credentials from DATABASE_URL ───────────────
# DATABASE_URL format: mysql+pymysql://user:password@host:port/dbname
def parse_db_url():
    url = DATABASE_URL
    # Remove dialect prefix
    url = url.replace("mysql+pymysql://", "")
    # Split user:pass@host:port/dbname
    user_pass, rest   = url.split("@")
    username, password = user_pass.split(":")
    host_port_db      = rest
    host_port, dbname = host_port_db.split("/")

    if ":" in host_port:
        host, port = host_port.split(":")
    else:
        host, port = host_port, "3306"

    return {
        "username": username,
        "password": password,
        "host":     host,
        "port":     port,
        "dbname":   dbname,
    }


# ════════════════════════════════════════════════════════════════
# SQL BACKUP
# Runs mysqldump and streams the output as a .sql file download
# ════════════════════════════════════════════════════════════════
@router.get("/sql")
def download_sql_backup(
    current_user: models.User = Depends(require_admin)
):
    try:
        creds = parse_db_url()

        # Build the mysqldump command
        # mysqldump must be in PATH (same bin folder as mysql.exe)
        cmd = [
            "mysqldump",
            f"--host={creds['host']}",
            f"--port={creds['port']}",
            f"--user={creds['username']}",
            f"--password={creds['password']}",
            "--single-transaction",   # consistent snapshot without locking
            "--routines",             # include stored procedures
            "--triggers",             # include triggers
            creds["dbname"],          # database name
        ]

        # Run mysqldump and capture output
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
        )

        if result.returncode != 0:
            raise HTTPException(
                status_code=500,
                detail=f"mysqldump failed: {result.stderr}"
            )

        # Prepare the SQL content as a downloadable stream
        sql_content = result.stdout
        date_str    = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename    = f"backup_{creds['dbname']}_{date_str}.sql"

        return StreamingResponse(
            io.StringIO(sql_content),
            media_type="application/sql",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except FileNotFoundError:
        raise HTTPException(
            status_code=500,
            detail="mysqldump not found. Make sure MySQL bin folder is in your system PATH."
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ════════════════════════════════════════════════════════════════
# EXCEL BACKUP
# Exports all tables to a multi-sheet Excel file
# ════════════════════════════════════════════════════════════════
@router.get("/excel")
def download_excel_backup(
    db:           Session     = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    try:
        wb = openpyxl.Workbook()
        # Remove default empty sheet
        wb.remove(wb.active)

        date_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        # ── Helper to style header rows ──
        def style_header(ws, headers, col_widths):
            ws.append(headers)
            for col_idx, cell in enumerate(ws[1], 1):
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="2563EB")
                cell.alignment = Alignment(horizontal="center")
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(i)
                ].width = width

        # ── Sheet 1: Products ──
        ws_products = wb.create_sheet("Products")
        style_header(ws_products,
            ["ID", "Name", "SKU", "Category", "Supplier",
             "Selling Price", "Cost Price", "Quantity",
             "Low Stock At", "Description", "Created At"],
            [5, 25, 12, 15, 15, 12, 12, 10, 10, 30, 18]
        )
        products = db.query(models.Product).all()
        for p in products:
            ws_products.append([
                p.id,
                p.name,
                p.sku or "",
                p.category.name  if p.category  else "",
                p.supplier.name  if p.supplier  else "",
                float(p.price),
                float(p.cost_price) if p.cost_price else 0.0,
                p.quantity,
                p.low_stock_threshold,
                p.description or "",
                p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
            ])

        # ── Sheet 2: Categories ──
        ws_cats = wb.create_sheet("Categories")
        style_header(ws_cats,
            ["ID", "Name", "Description", "Created At"],
            [5, 20, 40, 18]
        )
        categories = db.query(models.Category).all()
        for c in categories:
            ws_cats.append([
                c.id, c.name, c.description or "",
                c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
            ])

        # ── Sheet 3: Suppliers ──
        ws_sups = wb.create_sheet("Suppliers")
        style_header(ws_sups,
            ["ID", "Name", "Email", "Phone", "Address", "Created At"],
            [5, 20, 25, 15, 30, 18]
        )
        suppliers = db.query(models.Supplier).all()
        for s in suppliers:
            ws_sups.append([
                s.id, s.name, s.email or "",
                s.phone or "", s.address or "",
                s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "",
            ])

        # ── Sheet 4: Transactions ──
        ws_txns = wb.create_sheet("Transactions")
        style_header(ws_txns,
            ["ID", "Type", "Product", "Quantity", "Note", "Date"],
            [5, 8, 25, 10, 30, 18]
        )
        transactions = db.query(models.Transaction).order_by(
            models.Transaction.created_at.desc()
        ).all()
        for t in transactions:
            ws_txns.append([
                t.id,
                t.transaction_type,
                t.product.name if t.product else "",
                t.quantity,
                t.note or "",
                t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
            ])

        # ── Sheet 5: Users ──
        ws_users = wb.create_sheet("Users")
        style_header(ws_users,
            ["ID", "Username", "Role", "Active", "Created At"],
            [5, 20, 10, 8, 18]
        )
        users = db.query(models.User).all()
        for u in users:
            ws_users.append([
                u.id, u.username, u.role,
                "Yes" if u.is_active else "No",
                u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "",
            ])

        # ── Sheet 6: Activity Logs ──
        ws_logs = wb.create_sheet("Activity Logs")
        style_header(ws_logs,
            ["ID", "Action", "Detail", "User", "Timestamp"],
            [5, 25, 40, 15, 18]
        )
        logs = db.query(models.ActivityLog).order_by(
            models.ActivityLog.timestamp.desc()
        ).limit(1000).all()
        for l in logs:
            ws_logs.append([
                l.id, l.action, l.detail or "",
                l.user.username if l.user else "",
                l.timestamp.strftime("%Y-%m-%d %H:%M") if l.timestamp else "",
            ])

        # Save workbook to bytes buffer
        buffer   = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        date_file = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"backup_{date_file}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ════════════════════════════════════════════════════════════════
# IMPORT PRODUCTS FROM EXCEL
# Reads uploaded Excel file and creates products in bulk
# ════════════════════════════════════════════════════════════════
from fastapi import UploadFile, File

@router.post("/import/products")
async def import_products(
    file:         UploadFile        = File(...),
    db:           Session           = Depends(get_db),
    current_user: models.User       = Depends(require_admin),
):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx, .xls, or .csv files are supported"
        )

    try:
        content = await file.read()
        buffer  = io.BytesIO(content)
        wb      = openpyxl.load_workbook(buffer)
        ws      = wb.active  # Read from first sheet

        rows        = list(ws.iter_rows(values_only=True))
        header_row  = rows[0] if rows else []
        data_rows   = rows[1:] if len(rows) > 1 else []

        if not data_rows:
            raise HTTPException(status_code=400, detail="File is empty")

        # Map column headers to indices
        # Case-insensitive matching
        headers = {str(h).lower().strip(): i for i, h in enumerate(header_row) if h}

        created  = 0
        skipped  = 0
        errors   = []

        for row_num, row in enumerate(data_rows, start=2):
            try:
                # Get name — required
                name_idx = headers.get('name', 0)
                name     = str(row[name_idx]).strip() if row[name_idx] else None

                if not name or name == 'None':
                    skipped += 1
                    continue

                # Check if product with same name exists
                existing = db.query(models.Product).filter(
                    models.Product.name == name
                ).first()
                if existing:
                    skipped += 1
                    errors.append(f"Row {row_num}: '{name}' already exists — skipped")
                    continue

                # Get optional fields safely
                def get_val(key, default=None):
                    idx = headers.get(key)
                    if idx is None: return default
                    val = row[idx] if idx < len(row) else None
                    return val if val not in (None, '', 'None') else default

                # Get or create category
                category_id = None
                cat_name    = get_val('category')
                if cat_name:
                    cat = db.query(models.Category).filter(
                        models.Category.name == str(cat_name)
                    ).first()
                    if cat:
                        category_id = cat.id

                # Get or create supplier
                supplier_id = None
                sup_name    = get_val('supplier')
                if sup_name:
                    sup = db.query(models.Supplier).filter(
                        models.Supplier.name == str(sup_name)
                    ).first()
                    if sup:
                        supplier_id = sup.id

                # Parse numeric values safely
                def parse_float(val, default=0.0):
                    try: return float(val) if val else default
                    except: return default

                def parse_int(val, default=0):
                    try: return int(float(val)) if val else default
                    except: return default

                new_product = models.Product(
                    name                = name,
                    description         = str(get_val('description', '')),
                    price               = parse_float(get_val('price', get_val('selling price', 0))),
                    cost_price          = parse_float(get_val('cost price', get_val('cost_price', 0))),
                    quantity            = parse_int(get_val('quantity', 0)),
                    low_stock_threshold = parse_int(get_val('low stock at', get_val('low_stock_threshold', 10))),
                    sku                 = str(get_val('sku', '')).strip() or None,
                    category_id         = category_id,
                    supplier_id         = supplier_id,
                )

                db.add(new_product)
                created += 1

            except Exception as row_error:
                errors.append(f"Row {row_num}: {str(row_error)}")
                skipped += 1

        db.commit()

        log_activity(
            db, current_user.id,
            "Imported Products",
            f"Created {created}, skipped {skipped} from {file.filename}"
        )

        return {
            "message":  f"Import complete",
            "created":  created,
            "skipped":  skipped,
            "errors":   errors[:10],  # Return first 10 errors only
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ════════════════════════════════════════════════════════════════
# IMPORT TRANSACTIONS FROM EXCEL
# ════════════════════════════════════════════════════════════════
@router.post("/import/transactions")
async def import_transactions(
    file:         UploadFile        = File(...),
    db:           Session           = Depends(get_db),
    current_user: models.User       = Depends(require_admin),
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx or .xls files are supported"
        )

    try:
        content = await file.read()
        buffer  = io.BytesIO(content)
        wb      = openpyxl.load_workbook(buffer)
        ws      = wb.active

        rows       = list(ws.iter_rows(values_only=True))
        header_row = rows[0] if rows else []
        data_rows  = rows[1:] if len(rows) > 1 else []

        if not data_rows:
            raise HTTPException(status_code=400, detail="File is empty")

        headers = {str(h).lower().strip(): i for i, h in enumerate(header_row) if h}

        created = 0
        skipped = 0
        errors  = []

        for row_num, row in enumerate(data_rows, start=2):
            try:
                def get_val(key, default=None):
                    idx = headers.get(key)
                    if idx is None: return default
                    val = row[idx] if idx < len(row) else None
                    return val if val not in (None, '', 'None') else default

                # Get product by name
                product_name = get_val('product')
                if not product_name:
                    skipped += 1
                    errors.append(f"Row {row_num}: No product name — skipped")
                    continue

                product = db.query(models.Product).filter(
                    models.Product.name == str(product_name)
                ).first()

                if not product:
                    skipped += 1
                    errors.append(f"Row {row_num}: Product '{product_name}' not found — skipped")
                    continue

                # Get transaction type
                txn_type = str(get_val('type', '')).upper().strip()
                if txn_type not in ('IN', 'OUT'):
                    skipped += 1
                    errors.append(f"Row {row_num}: Type must be IN or OUT — skipped")
                    continue

                # Get quantity
                try:
                    qty = int(float(get_val('quantity', 0)))
                except:
                    skipped += 1
                    errors.append(f"Row {row_num}: Invalid quantity — skipped")
                    continue

                if qty <= 0:
                    skipped += 1
                    errors.append(f"Row {row_num}: Quantity must be > 0 — skipped")
                    continue

                # Check stock for OUT
                if txn_type == 'OUT' and product.quantity < qty:
                    skipped += 1
                    errors.append(f"Row {row_num}: Insufficient stock for '{product_name}' — skipped")
                    continue

                # Update product quantity
                if txn_type == 'IN':
                    product.quantity += qty
                else:
                    product.quantity -= qty

                # Create transaction
                new_txn = models.Transaction(
                    transaction_type = txn_type,
                    quantity         = qty,
                    note             = str(get_val('note', '')) or None,
                    product_id       = product.id,
                )
                db.add(new_txn)
                created += 1

            except Exception as row_error:
                errors.append(f"Row {row_num}: {str(row_error)}")
                skipped += 1

        db.commit()

        log_activity(
            db, current_user.id,
            "Imported Transactions",
            f"Created {created}, skipped {skipped} from {file.filename}"
        )

        return {
            "message": "Import complete",
            "created": created,
            "skipped": skipped,
            "errors":  errors[:10],
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── DOWNLOAD TEMPLATE FILES ──────────────────────────────────────
# Gives users a pre-formatted Excel template to fill in

@router.get("/template/products")
def download_products_template(
    current_user: models.User = Depends(get_current_user)
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["name", "price", "cost_price", "quantity",
               "low_stock_threshold", "sku", "category",
               "supplier", "description"]

    # Style headers
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    # Add example row
    ws.append([
        "iPhone 15", 999.99, 750.00, 50, 10,
        "ELEC-001", "Electronics", "Tech Supplier", "Latest iPhone"
    ])

    # Set column widths
    widths = [20, 10, 10, 10, 15, 12, 15, 15, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_template.xlsx"}
    )


@router.get("/template/transactions")
def download_transactions_template(
    current_user: models.User = Depends(get_current_user)
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["product", "type", "quantity", "note"]

    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    # Example rows
    ws.append(["iPhone 15", "IN",  50, "Restocked from supplier"])
    ws.append(["iPhone 15", "OUT", 5,  "Sold to customer"])

    widths = [20, 8, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=transactions_template.xlsx"}
    )